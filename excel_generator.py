"""
Excel Dashboard Generator
Create Excel files with pivot tables and charts
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelDashboard:
    """Generate Excel dashboards"""
    
    def __init__(self, data, output_dir='excel'):
        self.df = data.copy()
        self.output_dir = output_dir
    
    def create_kpi_dashboard(self, kpis, filename='financial_dashboard.xlsx'):
        """
        Create main KPI dashboard
        Simple and clean layout
        """
        
        print("Creating KPI Dashboard...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "KPI Dashboard"
        
        # Title
        ws['A1'] = "Financial Performance Dashboard"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # KPI Cards
        kpi_row = 3
        
        # Define KPIs to display
        kpi_list = [
            ('Total Revenue', f"₹{kpis['total_revenue']:,.2f}"),
            ('Total Profit', f"₹{kpis['total_profit']:,.2f}"),
            ('Profit Margin', f"{kpis['avg_profit_margin']:.2f}%"),
            ('Total Transactions', f"{kpis['total_transactions']:,}"),
            ('Avg Transaction Value', f"₹{kpis['avg_transaction_value']:,.2f}"),
            ('Total Customers', f"{kpis['total_customers']:,}")
        ]
        
        for i, (label, value) in enumerate(kpi_list):
            row = kpi_row + i * 2
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = Font(size=12)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        
        # Save workbook
        filepath = f"{self.output_dir}/{filename}"
        wb.save(filepath)
        print(f"✓ Saved: {filepath}")
        
        return filepath
    
    def create_segment_analysis(self, segment_data, filename='segment_analysis.xlsx'):
        """
        Create segment performance analysis
        With simple pivot table style layout
        """
        
        print("Creating Segment Analysis...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Segment Performance"
        
        # Title
        ws['A1'] = "Segment Performance Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:G1')
        
        # Headers
        headers = ['Segment', 'Revenue', 'Profit', 'Margin %', 'Transactions', 'Avg Revenue']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        # Data
        segment_summary = self.df.groupby('segment').agg({
            'revenue': ['sum', 'mean'],
            'profit': 'sum',
            'profit_margin': 'mean',
            'transaction_id': 'count'
        }).round(2)
        
        row_num = 4
        for segment in segment_summary.index:
            ws.cell(row=row_num, column=1).value = segment
            ws.cell(row=row_num, column=2).value = segment_summary.loc[segment, ('revenue', 'sum')]
            ws.cell(row=row_num, column=3).value = segment_summary.loc[segment, ('profit', 'sum')]
            ws.cell(row=row_num, column=4).value = segment_summary.loc[segment, ('profit_margin', 'mean')]
            ws.cell(row=row_num, column=5).value = segment_summary.loc[segment, ('transaction_id', 'count')]
            ws.cell(row=row_num, column=6).value = segment_summary.loc[segment, ('revenue', 'mean')]
            row_num += 1
        
        # Format numbers
        for row in range(4, row_num):
            ws.cell(row=row, column=2).number_format = '₹#,##0.00'
            ws.cell(row=row, column=3).number_format = '₹#,##0.00'
            ws.cell(row=row, column=4).number_format = '0.00'
            ws.cell(row=row, column=6).number_format = '₹#,##0.00'
        
        # Adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[chr(64+col)].width = 15
        
        filepath = f"{self.output_dir}/{filename}"
        wb.save(filepath)
        print(f"✓ Saved: {filepath}")
        
        return filepath
    
    def create_monthly_trends(self, filename='monthly_trends.xlsx'):
        """Create monthly revenue and profit trends"""
        
        print("Creating Monthly Trends...")
        
        self.df['date'] = pd.to_datetime(self.df['date'])
        
        monthly = self.df.groupby([self.df['date'].dt.year, 
                                   self.df['date'].dt.month]).agg({
            'revenue': 'sum',
            'profit': 'sum',
            'profit_margin': 'mean'
        }).round(2).reset_index()
        
        monthly.columns = ['Year', 'Month', 'Revenue', 'Profit', 'Margin']
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Monthly Trends"
        
        # Title
        ws['A1'] = "Monthly Revenue & Profit Trends"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Write data
        for r in dataframe_to_rows(monthly, index=False, header=True):
            ws.append(r)
        
        # Format headers
        for cell in ws[3]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        # Format numbers
        for row in range(4, len(monthly) + 4):
            ws.cell(row=row, column=3).number_format = '₹#,##0.00'
            ws.cell(row=row, column=4).number_format = '₹#,##0.00'
            ws.cell(row=row, column=5).number_format = '0.00'
        
        filepath = f"{self.output_dir}/{filename}"
        wb.save(filepath)
        print(f"✓ Saved: {filepath}")
        
        return filepath
    
    def create_forecast_report(self, forecast_data, accuracy, filename='forecast_report.xlsx'):
        """Create revenue forecast report"""
        
        print("Creating Forecast Report...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Revenue Forecast"
        
        # Title
        ws['A1'] = "6-Month Revenue Forecast"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Forecast accuracy
        ws['A3'] = "Forecast Accuracy (MAPE):"
        ws['A3'].font = Font(bold=True)
        ws['B3'] = f"{accuracy['mape']:.2f}%"
        ws['B3'].font = Font(size=12, color='00B050' if accuracy['mape'] < 10 else '000000')
        
        # Headers
        headers = ['Date', 'Forecasted Revenue', 'Lower Bound', 'Upper Bound']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        # Forecast data
        row_num = 6
        for _, row in forecast_data.iterrows():
            ws.cell(row=row_num, column=1).value = row['date'].strftime('%Y-%m')
            ws.cell(row=row_num, column=2).value = row['final_forecast']
            ws.cell(row=row_num, column=3).value = row['lower_bound']
            ws.cell(row=row_num, column=4).value = row['upper_bound']
            
            # Format as currency
            for col in range(2, 5):
                ws.cell(row=row_num, column=col).number_format = '₹#,##0.00'
            
            row_num += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D']:
            ws.column_dimensions[col].width = 18
        
        filepath = f"{self.output_dir}/{filename}"
        wb.save(filepath)
        print(f"✓ Saved: {filepath}")
        
        return filepath
